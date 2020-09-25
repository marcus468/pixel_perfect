import xlrd
import os
import os.path
class gen_excel_report:
    global template_path, df
    def main(self):
        
        cell_list,range_list,max_record= self.read_template(self.template_path)
        self.gen_report(cell_list,range_list,self.df,max_record,self.template_path)
    def read_template(self,path):
        import pandas as pd
        import re

        workbook = xlrd.open_workbook(path)
        sheet_names = workbook.sheet_names()
        cell_list = []
        range_list = []
        record_list = []
        xl_sheet = workbook.sheet_by_name(sheet_names[0])
        for name in workbook.name_map:
            dic ={}
            cell_obj = workbook.name_and_scope_map.get((name, -1))
            dic["variable"] = name
            dic["position"] = cell_obj.formula_text
            if re.search(":",cell_obj.formula_text) is not None:
                dic["variable"] = name
                dic["position"] = cell_obj.formula_text
                range_list.append(dic)

                record_list.append(int(cell_obj.formula_text.split(":")[1].split("$")[2]) +1 - int(cell_obj.formula_text.split(":")[0].split("$")[2]))
            else:
                dic["variable"] = name
                dic["position"] = cell_obj.formula_text
                cell_list.append(dic)
        if len(record_list)>1:
            max_record = max(record_list)
        else:
            max_record = 0


        return cell_list,range_list,max_record

    def insert_list_to_excel(self,list_, from_col,from_row,to_col,to_row,wb):
        import openpyxl
        i = 0

        ws = wb.worksheets[0]
        row = int(from_row)-1
        sheet_num = 0

        for value in list_:
            row +=1

            if int(row) > int(to_row):
                row = int(from_row)
                sheet_num+=1
                i+=1
                ws = wb.worksheets[sheet_num]
            ws[str(from_col)+str(row)] = value
    def gen_report(self,cell_list,range_list,df,max_record,path):
        import math
        import openpyxl
        import re
        from copy import copy
        parent_path = os.path.dirname(path)
        os.system("cp %s %s" % (path,os.path.join(parent_path,"report.xlsx")))
        wb = openpyxl.load_workbook(os.path.join(parent_path,"report.xlsx"))

        ind=0

        for p in wb.get_sheet_names():
            ws = wb[p]
            for ele in cell_list:
                var = ele.get("variable")
                position = ele.get("position")

                position = ele.get("position").split("$")[1] + ele.get("position").split("$")[2]
                ws[position] = df[var].iloc[-1]

        if len(df)/max_record > 0:
            num_sheet = int(math.ceil(len(df)/max_record))-1

            for a in range(num_sheet):
                ind+=1
                source=wb.get_sheet_by_name(wb.get_sheet_names()[0])
                wb.copy_worksheet(source)
        for ele in range_list:
            var = ele.get("variable")
            position = ele.get("position")
            from_col = position.split(":")[0].split("$")[1]
            from_row = position.split(":")[0].split("$")[2]
            to_col = position.split(":")[1].split("$")[1]
            to_row = position.split(":")[1].split("$")[2]

            self.insert_list_to_excel(df[var],from_col,from_row,to_col,to_row,wb)
        ind=0

        for name in wb.get_sheet_names():
            ind+=1
            ss_sheet = wb[name]
            ss_sheet.title = 'page'+str(ind)
        wb.save(path)

    def __init__(self, template_path,df):



        self.template_path=template_path
        self.df = df

import pandas as pd 

df = pd.read_excel("/home/marcus/code/python/pixel_perfect_study/pixel_perfect_sample_data_frame.xlsx")
df = df.fillna("")


gen = gen_excel_report("/home/marcus/code/python/pixel_perfect_study/template.xlsx",df)
gen.main()